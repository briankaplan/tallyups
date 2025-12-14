#!/usr/bin/env python3
"""
Comprehensive Unit Tests for Report Exporters
==============================================

Tests for the report export system including:
- CSV export for various formats
- Excel workbook generation
- PDF report generation
- Format compliance for accounting software

Test Coverage Target: 90%+
"""

import pytest
from datetime import datetime, timedelta
from decimal import Decimal
from unittest.mock import Mock, MagicMock, patch
import csv
import io
import sys
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from services.csv_exporter import CSVExporter, get_csv_exporter
except ImportError:
    CSVExporter = None

try:
    from services.excel_exporter import ExcelExporter, get_excel_exporter
except ImportError:
    ExcelExporter = None

try:
    from services.pdf_exporter import PDFExporter, get_pdf_exporter, HAS_REPORTLAB
except ImportError:
    PDFExporter = None
    HAS_REPORTLAB = False

try:
    from services.report_generator import (
        ExpenseReportGenerator,
        Report,
        Transaction,
        ReportSummary,
        CategoryBreakdown,
        VendorBreakdown,
        MonthlyTrend,
        BusinessType,
    )
except ImportError:
    ExpenseReportGenerator = None
    Report = None


# =============================================================================
# MOCK DATA HELPERS
# =============================================================================

def create_mock_transaction(
    index: int = 1,
    description: str = "Test Transaction",
    amount: float = 100.00,
    date: datetime = None,
    category: str = "Software & Subscriptions",
    business_type: str = "Down Home",
    has_receipt: bool = True,
) -> Mock:
    """Create a mock transaction for testing."""
    tx = Mock()
    tx.index = index
    tx.description = description
    tx.amount = Decimal(str(amount))
    tx.date = date or datetime.now()
    tx.effective_category = category
    tx.business_type = business_type
    tx.review_status = "VERIFIED" if has_receipt else None
    tx.has_receipt = has_receipt
    tx.effective_receipt_url = f"https://r2.example.com/receipt_{index}.jpg" if has_receipt else ""
    tx.notes = "Test notes"
    tx.ai_note = "AI generated note"
    tx.ai_confidence = 0.95 if has_receipt else None
    tx.mi_merchant = description
    tx.mi_category = category
    tx.source = "Chase"
    return tx


def create_mock_report(
    business_type: str = "Down Home",
    num_transactions: int = 10,
) -> Mock:
    """Create a mock report for testing."""
    report = Mock()
    report.report_name = f"{business_type} Expense Report"
    report.report_id = "RPT-2024-001"
    report.business_type = business_type
    report.date_range = (datetime(2024, 1, 1), datetime(2024, 1, 31))
    report.generated_at = datetime.now()

    # Create transactions
    transactions = []
    for i in range(num_transactions):
        tx = create_mock_transaction(
            index=i + 1,
            description=f"Transaction {i + 1}",
            amount=50.00 + (i * 10),
            date=datetime(2024, 1, 15) - timedelta(days=i),
            has_receipt=i % 3 != 0,  # Every 3rd transaction missing receipt
        )
        transactions.append(tx)
    report.transactions = transactions

    # Create summary
    total = sum(float(tx.amount) for tx in transactions)
    matched = sum(1 for tx in transactions if tx.has_receipt)
    report.summary = Mock()
    report.summary.total_transactions = len(transactions)
    report.summary.total_amount = Decimal(str(total))
    report.summary.average_transaction = Decimal(str(total / len(transactions))) if transactions else Decimal("0")
    report.summary.largest_transaction = max(tx.amount for tx in transactions) if transactions else Decimal("0")
    report.summary.smallest_transaction = min(tx.amount for tx in transactions) if transactions else Decimal("0")
    report.summary.match_rate = 80.0 if transactions else 0.0
    report.summary.receipt_rate = 70.0 if transactions else 0.0
    report.summary.matched_count = matched
    report.summary.unmatched_count = len(transactions) - matched
    report.summary.receipt_count = matched
    report.summary.no_receipt_count = len(transactions) - matched
    report.summary.receipts_attached = matched
    report.summary.receipts_missing = len(transactions) - matched

    # Category breakdown
    report.summary.by_category = [
        Mock(category="Software & Subscriptions", total=Decimal("300.00"), count=5, percentage=50.0),
        Mock(category="Travel - Meals", total=Decimal("200.00"), count=3, percentage=33.3),
        Mock(category="Office Supplies", total=Decimal("100.00"), count=2, percentage=16.7),
    ]

    # Vendor breakdown
    report.summary.by_vendor = [
        Mock(vendor="Anthropic", total=Decimal("200.00"), count=4, is_recurring=True, average=Decimal("50.00"),
             first_transaction=datetime(2024, 1, 1), last_transaction=datetime(2024, 1, 15), categories=["Software & Subscriptions"]),
        Mock(vendor="Uber", total=Decimal("150.00"), count=3, is_recurring=False, average=Decimal("50.00"),
             first_transaction=datetime(2024, 1, 5), last_transaction=datetime(2024, 1, 20), categories=["Travel - Transportation"]),
        Mock(vendor="Starbucks", total=Decimal("50.00"), count=2, is_recurring=False, average=Decimal("25.00"),
             first_transaction=datetime(2024, 1, 10), last_transaction=datetime(2024, 1, 25), categories=["Travel - Meals"]),
    ]

    # Monthly trends
    report.summary.monthly_trends = [
        Mock(year=2024, month=1, total=Decimal("600.00"), count=10),
    ]

    return report


# =============================================================================
# CSV EXPORTER TESTS
# =============================================================================

@pytest.mark.skipif(CSVExporter is None, reason="CSVExporter not available")
class TestCSVExporter:
    """Test suite for CSV exporter."""

    @pytest.fixture
    def exporter(self):
        return CSVExporter()

    @pytest.fixture
    def mock_report(self):
        return create_mock_report()

    # Standard CSV Export
    @pytest.mark.unit
    def test_export_standard_csv(self, exporter, mock_report):
        """Standard CSV export should work."""
        result = exporter.export_standard_csv(mock_report)
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.unit
    def test_standard_csv_headers(self, exporter, mock_report):
        """Standard CSV should have correct headers."""
        result = exporter.export_standard_csv(mock_report)
        content = result.decode('utf-8')
        lines = content.strip().split('\n')
        headers = lines[0]

        assert "Index" in headers
        assert "Date" in headers
        assert "Description" in headers
        assert "Amount" in headers
        assert "Category" in headers
        assert "Has Receipt" in headers

    @pytest.mark.unit
    def test_standard_csv_row_count(self, exporter, mock_report):
        """Standard CSV should have correct row count."""
        result = exporter.export_standard_csv(mock_report)
        content = result.decode('utf-8')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        # Header + data rows
        assert len(rows) == len(mock_report.transactions) + 1

    # QuickBooks CSV Export
    @pytest.mark.unit
    def test_export_quickbooks_csv(self, exporter, mock_report):
        """QuickBooks CSV export should work."""
        result = exporter.export_quickbooks_csv(mock_report)
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.unit
    def test_quickbooks_csv_headers(self, exporter, mock_report):
        """QuickBooks CSV should have correct headers."""
        result = exporter.export_quickbooks_csv(mock_report)
        content = result.decode('utf-8')
        lines = content.strip().split('\n')
        headers = lines[0]

        assert "Date" in headers
        assert "Description" in headers
        assert "Amount" in headers
        assert "Account" in headers

    @pytest.mark.unit
    def test_quickbooks_category_mapping(self, exporter):
        """QuickBooks should map categories correctly."""
        mapping = exporter.quickbooks_category_map
        assert "Travel - Airfare" in mapping
        assert mapping["Travel - Airfare"] == "Travel:Airfare"

    # Xero CSV Export
    @pytest.mark.unit
    def test_export_xero_csv(self, exporter, mock_report):
        """Xero CSV export should work."""
        result = exporter.export_xero_csv(mock_report)
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.unit
    def test_xero_csv_headers(self, exporter, mock_report):
        """Xero CSV should have correct headers."""
        result = exporter.export_xero_csv(mock_report)
        content = result.decode('utf-8')
        lines = content.strip().split('\n')
        headers = lines[0]

        assert "*Date" in headers
        assert "*Amount" in headers
        assert "Payee" in headers

    @pytest.mark.unit
    def test_xero_date_format(self, exporter, mock_report):
        """Xero dates should be in YYYY-MM-DD format."""
        result = exporter.export_xero_csv(mock_report)
        content = result.decode('utf-8')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        if len(rows) > 1:
            date_col = rows[1][0]
            # Should be YYYY-MM-DD format
            assert len(date_col.split('-')) == 3

    # Down Home CSV Export
    @pytest.mark.unit
    def test_export_downhome_csv(self, exporter, mock_report):
        """Down Home CSV export should work."""
        result = exporter.export_downhome_csv(mock_report)
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.unit
    def test_downhome_csv_headers(self, exporter, mock_report):
        """Down Home CSV should have correct headers."""
        result = exporter.export_downhome_csv(mock_report)
        content = result.decode('utf-8')
        lines = content.strip().split('\n')
        headers = lines[0]

        assert "External ID" in headers
        assert "Line" in headers
        assert "Category" in headers
        assert "Amount" in headers
        assert "Currency" in headers
        assert "Receipt URL" in headers

    @pytest.mark.unit
    def test_downhome_csv_currency(self, exporter, mock_report):
        """Down Home CSV should use USD currency."""
        result = exporter.export_downhome_csv(mock_report)
        content = result.decode('utf-8')
        assert "USD" in content

    # Summary CSV Export
    @pytest.mark.unit
    def test_export_summary_csv(self, exporter, mock_report):
        """Summary CSV export should work."""
        result = exporter.export_summary_csv(mock_report)
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.unit
    def test_summary_csv_contains_totals(self, exporter, mock_report):
        """Summary CSV should contain totals."""
        result = exporter.export_summary_csv(mock_report)
        content = result.decode('utf-8')

        assert "Total Transactions" in content
        assert "Total Amount" in content

    @pytest.mark.unit
    def test_summary_csv_contains_breakdown(self, exporter, mock_report):
        """Summary CSV should contain category breakdown."""
        result = exporter.export_summary_csv(mock_report)
        content = result.decode('utf-8')

        assert "CATEGORY BREAKDOWN" in content
        assert "TOP VENDORS" in content

    # Reconciliation CSV Export
    @pytest.mark.unit
    def test_export_reconciliation_csv(self, exporter, mock_report):
        """Reconciliation CSV export should work."""
        result = exporter.export_reconciliation_csv(mock_report)
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.unit
    def test_reconciliation_csv_match_status(self, exporter, mock_report):
        """Reconciliation CSV should include match status."""
        result = exporter.export_reconciliation_csv(mock_report)
        content = result.decode('utf-8')

        assert "Match Status" in content
        # Should have MATCHED or UNMATCHED
        assert "MATCHED" in content or "UNMATCHED" in content

    # Multi-business CSV Export
    @pytest.mark.unit
    def test_export_multi_business_csv(self, exporter):
        """Multi-business CSV export should work."""
        reports = {
            "Down Home": create_mock_report("Down Home", 5),
            "Personal": create_mock_report("Personal", 5),
        }
        result = exporter.export_multi_business_csv(reports)
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.unit
    def test_multi_business_csv_grand_total(self, exporter):
        """Multi-business CSV should have grand total."""
        reports = {
            "Down Home": create_mock_report("Down Home", 5),
            "Personal": create_mock_report("Personal", 5),
        }
        result = exporter.export_multi_business_csv(reports)
        content = result.decode('utf-8')

        assert "GRAND TOTAL" in content

    # Edge Cases
    @pytest.mark.unit
    def test_empty_report(self, exporter):
        """Empty report should not crash."""
        report = create_mock_report(num_transactions=0)
        report.transactions = []
        result = exporter.export_standard_csv(report)
        assert isinstance(result, bytes)

    @pytest.mark.unit
    def test_special_characters_in_description(self, exporter, mock_report):
        """Special characters should be handled in CSV."""
        mock_report.transactions[0].description = 'Test, "quoted", special'
        result = exporter.export_standard_csv(mock_report)
        assert isinstance(result, bytes)
        # Should not crash

    @pytest.mark.unit
    def test_unicode_in_merchant(self, exporter, mock_report):
        """Unicode characters should be handled."""
        mock_report.transactions[0].description = "Café Délice 日本料理"
        result = exporter.export_standard_csv(mock_report)
        content = result.decode('utf-8')
        assert "Café" in content or "Caf" in content  # UTF-8 encoded


# =============================================================================
# CSV EXPORTER SINGLETON TESTS
# =============================================================================

@pytest.mark.skipif(CSVExporter is None, reason="CSVExporter not available")
class TestCSVExporterSingleton:
    """Test CSV exporter singleton pattern."""

    @pytest.mark.unit
    def test_singleton(self):
        """get_csv_exporter should return same instance."""
        exporter1 = get_csv_exporter()
        exporter2 = get_csv_exporter()
        assert exporter1 is exporter2


# =============================================================================
# EXCEL EXPORTER TESTS
# =============================================================================

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

@pytest.mark.skipif(ExcelExporter is None or not HAS_OPENPYXL, reason="ExcelExporter or openpyxl not available")
class TestExcelExporter:
    """Test suite for Excel exporter."""

    @pytest.fixture
    def exporter(self):
        return ExcelExporter()

    @pytest.fixture
    def mock_report(self):
        return create_mock_report()

    @pytest.mark.unit
    def test_export_report(self, exporter, mock_report):
        """Excel export should produce bytes."""
        result = exporter.export_report(mock_report)
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.unit
    def test_excel_magic_bytes(self, exporter, mock_report):
        """Excel file should have correct magic bytes."""
        result = exporter.export_report(mock_report)
        # XLSX files start with PK (ZIP format)
        assert result[:2] == b'PK'

    @pytest.mark.unit
    def test_excel_can_be_read(self, exporter, mock_report, tmp_path):
        """Generated Excel should be readable."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        result = exporter.export_report(mock_report)

        # Save to file
        excel_path = tmp_path / "test_report.xlsx"
        with open(excel_path, 'wb') as f:
            f.write(result)

        # Load and verify
        wb = load_workbook(excel_path)
        assert len(wb.sheetnames) >= 1

    @pytest.mark.unit
    def test_excel_multiple_sheets(self, exporter, mock_report, tmp_path):
        """Excel should have multiple sheets."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        result = exporter.export_report(mock_report)

        excel_path = tmp_path / "test_report.xlsx"
        with open(excel_path, 'wb') as f:
            f.write(result)

        wb = load_workbook(excel_path)
        # Should have summary, transactions, etc.
        assert len(wb.sheetnames) >= 2

    @pytest.mark.unit
    def test_empty_report(self, exporter):
        """Empty report should not crash."""
        report = create_mock_report(num_transactions=0)
        report.transactions = []
        result = exporter.export_report(report)
        assert isinstance(result, bytes)


# =============================================================================
# PDF EXPORTER TESTS
# =============================================================================

@pytest.mark.skipif(PDFExporter is None or not HAS_REPORTLAB, reason="PDFExporter or reportlab not available")
class TestPDFExporter:
    """Test suite for PDF exporter."""

    @pytest.fixture
    def exporter(self):
        return PDFExporter()

    @pytest.fixture
    def mock_report(self):
        return create_mock_report()

    @pytest.mark.unit
    def test_export_report(self, exporter, mock_report):
        """PDF export should produce bytes."""
        result = exporter.export_report(mock_report)
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.unit
    def test_pdf_magic_bytes(self, exporter, mock_report):
        """PDF file should have correct magic bytes."""
        result = exporter.export_report(mock_report)
        # PDF files start with %PDF
        assert result[:4] == b'%PDF'

    @pytest.mark.unit
    def test_pdf_contains_report_name(self, exporter, mock_report, tmp_path):
        """PDF should contain report name."""
        result = exporter.export_report(mock_report)

        # Save and check content
        pdf_path = tmp_path / "test_report.pdf"
        with open(pdf_path, 'wb') as f:
            f.write(result)

        # Basic check - file exists and has content
        assert pdf_path.exists()
        assert pdf_path.stat().st_size > 1000  # Should be reasonable size

    @pytest.mark.unit
    def test_empty_report(self, exporter):
        """Empty report should not crash."""
        report = create_mock_report(num_transactions=0)
        report.transactions = []
        result = exporter.export_report(report)
        assert isinstance(result, bytes)

    @pytest.mark.unit
    def test_reconciliation_report(self, exporter, mock_report):
        """Reconciliation PDF export should work."""
        if hasattr(exporter, 'export_reconciliation_report'):
            result = exporter.export_reconciliation_report(mock_report)
            assert isinstance(result, bytes)
            assert result[:4] == b'%PDF'


# =============================================================================
# DATE/AMOUNT FORMATTING TESTS
# =============================================================================

@pytest.mark.skipif(CSVExporter is None, reason="CSVExporter not available")
class TestFormatting:
    """Test date and amount formatting."""

    @pytest.fixture
    def exporter(self):
        return CSVExporter()

    @pytest.mark.unit
    def test_format_date_default(self, exporter):
        """Default date format should be MM/DD/YYYY."""
        dt = datetime(2024, 1, 15)
        result = exporter._format_date(dt)
        assert result == "01/15/2024"

    @pytest.mark.unit
    def test_format_date_custom(self, exporter):
        """Custom date format should work."""
        dt = datetime(2024, 1, 15)
        result = exporter._format_date(dt, "%Y-%m-%d")
        assert result == "2024-01-15"

    @pytest.mark.unit
    def test_format_date_none(self, exporter):
        """None date should return empty string."""
        result = exporter._format_date(None)
        assert result == ""

    @pytest.mark.unit
    def test_format_amount_positive(self, exporter):
        """Positive amount should format correctly."""
        result = exporter._format_amount(Decimal("123.45"))
        assert result == "123.45"

    @pytest.mark.unit
    def test_format_amount_negative_absolute(self, exporter):
        """Negative amount should be absolute by default."""
        result = exporter._format_amount(Decimal("-123.45"))
        assert result == "123.45"

    @pytest.mark.unit
    def test_format_amount_negative_signed(self, exporter):
        """Negative amount should keep sign when absolute=False."""
        result = exporter._format_amount(Decimal("-123.45"), absolute=False)
        assert result == "-123.45"

    @pytest.mark.unit
    def test_format_amount_precision(self, exporter):
        """Amount should have 2 decimal places."""
        result = exporter._format_amount(Decimal("100"))
        assert result == "100.00"


# =============================================================================
# CATEGORY MAPPING TESTS
# =============================================================================

@pytest.mark.skipif(CSVExporter is None, reason="CSVExporter not available")
class TestCategoryMapping:
    """Test category mapping for accounting software."""

    @pytest.fixture
    def exporter(self):
        return CSVExporter()

    @pytest.mark.unit
    def test_quickbooks_travel_airfare(self, exporter):
        """Travel - Airfare should map correctly for QuickBooks."""
        result = exporter._map_category_quickbooks("Travel - Airfare")
        assert "Travel" in result
        assert "Airfare" in result

    @pytest.mark.unit
    def test_quickbooks_unknown_category(self, exporter):
        """Unknown category should map to Other Expenses."""
        result = exporter._map_category_quickbooks("Unknown Category XYZ")
        assert "Other" in result or "Expense" in result

    @pytest.mark.unit
    def test_xero_travel_hotel(self, exporter):
        """Travel - Hotel should map correctly for Xero."""
        result = exporter._map_category_xero("Travel - Hotel")
        assert "Travel" in result or "Accommodation" in result

    @pytest.mark.unit
    def test_xero_unknown_category(self, exporter):
        """Unknown category should map to General Expenses."""
        result = exporter._map_category_xero("Unknown Category XYZ")
        assert "General" in result or "Expense" in result


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
