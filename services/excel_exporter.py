"""
Excel Expense Report Exporter
=============================

Generates professional Excel workbooks with:
- Summary sheet with totals and charts
- Detail sheets per business type
- Category breakdown with charts
- Vendor analysis
- Pivot table data
- Conditional formatting
"""

import io
import logging
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

logger = logging.getLogger(__name__)

# Try to import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed - Excel export will be limited")


# Style constants
BRAND_GREEN = "00FF88"
DARK_BG = "0B0D10"
PANEL_BG = "111111"
HEADER_BG = "1A1A1A"
TEXT_COLOR = "F0F0F0"
MUTED_COLOR = "888888"
SUCCESS_COLOR = "00FF88"
WARNING_COLOR = "FFD85E"
ERROR_COLOR = "FF4E6A"


class ExcelExporter:
    """
    Generate professional Excel expense reports.
    """

    def __init__(self):
        """Initialize Excel exporter."""
        self.styles_created = False

        if not HAS_OPENPYXL:
            logger.warning("openpyxl not available - install with: pip install openpyxl")

    def _create_styles(self, wb: 'Workbook'):
        """Create named styles for the workbook."""
        if self.styles_created:
            return

        # Header style
        header_style = NamedStyle(name='header_style')
        header_style.font = Font(bold=True, color=TEXT_COLOR, size=11)
        header_style.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_style.border = Border(
            bottom=Side(style='thin', color=BRAND_GREEN)
        )

        # Currency style
        currency_style = NamedStyle(name='currency_style')
        currency_style.number_format = '$#,##0.00'
        currency_style.alignment = Alignment(horizontal='right')

        # Percentage style
        percent_style = NamedStyle(name='percent_style')
        percent_style.number_format = '0.0%'
        percent_style.alignment = Alignment(horizontal='center')

        # Date style
        date_style = NamedStyle(name='date_style')
        date_style.number_format = 'MM/DD/YYYY'
        date_style.alignment = Alignment(horizontal='center')

        # Title style
        title_style = NamedStyle(name='title_style')
        title_style.font = Font(bold=True, color=BRAND_GREEN, size=16)
        title_style.alignment = Alignment(horizontal='left')

        # Add styles to workbook
        try:
            wb.add_named_style(header_style)
            wb.add_named_style(currency_style)
            wb.add_named_style(percent_style)
            wb.add_named_style(date_style)
            wb.add_named_style(title_style)
            self.styles_created = True
        except ValueError:
            # Styles already exist
            pass

    def _apply_header_style(self, ws, row: int, start_col: int = 1, end_col: int = None):
        """Apply header styling to a row."""
        end_col = end_col or ws.max_column
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color=TEXT_COLOR, size=11)
            cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(bottom=Side(style='thin', color=BRAND_GREEN))

    def _auto_column_width(self, ws, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths based on content."""
        for col in ws.columns:
            max_length = 0
            # Handle MergedCell objects that don't have column_letter attribute
            column = None
            for cell in col:
                if hasattr(cell, 'column_letter'):
                    column = cell.column_letter
                    break

            if column is None:
                continue  # Skip if we can't determine the column

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width

    def export_report(self, report: 'Report') -> bytes:
        """
        Export report to Excel workbook.

        Args:
            report: Report object to export

        Returns:
            Excel file as bytes
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")

        wb = Workbook()
        self.styles_created = False
        self._create_styles(wb)

        # Remove default sheet
        default_sheet = wb.active
        default_sheet.title = "Summary"

        # Create summary sheet
        self._create_summary_sheet(wb, report)

        # Create transactions sheet
        self._create_transactions_sheet(wb, report)

        # Create category breakdown sheet
        self._create_category_sheet(wb, report)

        # Create vendor analysis sheet
        self._create_vendor_sheet(wb, report)

        # Create monthly trends sheet
        self._create_trends_sheet(wb, report)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _create_summary_sheet(self, wb: 'Workbook', report: 'Report'):
        """Create the summary sheet with key metrics."""
        ws = wb.active
        ws.title = "Summary"

        # Report header
        ws['A1'] = report.report_name
        ws['A1'].font = Font(bold=True, color=BRAND_GREEN, size=18)
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Generated: {report.generated_at.strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = Font(color=MUTED_COLOR, size=10)

        ws['A3'] = f"Date Range: {report.date_range[0].strftime('%m/%d/%Y')} - {report.date_range[1].strftime('%m/%d/%Y')}"
        ws['A3'].font = Font(color=MUTED_COLOR, size=10)

        ws['A4'] = f"Business Type: {report.business_type}"
        ws['A4'].font = Font(color=TEXT_COLOR, size=11)

        # Key metrics section
        row = 6
        ws.cell(row=row, column=1, value="KEY METRICS").font = Font(bold=True, color=BRAND_GREEN, size=14)
        ws.merge_cells(f'A{row}:C{row}')

        row += 2
        metrics = [
            ("Total Transactions", report.summary.total_transactions, None),
            ("Total Amount", float(report.summary.total_amount), 'currency_style'),
            ("Average Transaction", float(report.summary.average_transaction), 'currency_style'),
            ("Largest Transaction", float(report.summary.largest_transaction), 'currency_style'),
            ("", "", None),
            ("Matched Transactions", report.summary.matched_count, None),
            ("Unmatched Transactions", report.summary.unmatched_count, None),
            ("Match Rate", report.summary.match_rate / 100, 'percent_style'),
            ("", "", None),
            ("Receipts Attached", report.summary.receipts_attached, None),
            ("Receipts Missing", report.summary.receipts_missing, None),
            ("Receipt Rate", report.summary.receipt_rate / 100, 'percent_style'),
        ]

        for metric_name, metric_value, style in metrics:
            ws.cell(row=row, column=1, value=metric_name).font = Font(color=TEXT_COLOR)
            value_cell = ws.cell(row=row, column=2, value=metric_value)
            if style == 'currency_style':
                value_cell.number_format = '$#,##0.00'
            elif style == 'percent_style':
                value_cell.number_format = '0.0%'
            value_cell.font = Font(bold=True, color=BRAND_GREEN)
            row += 1

        # Top Categories section
        row += 2
        ws.cell(row=row, column=1, value="TOP CATEGORIES").font = Font(bold=True, color=BRAND_GREEN, size=14)
        row += 2

        headers = ["Category", "Amount", "Count", "% of Total"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, 4)

        row += 1
        for cat in report.summary.by_category[:10]:
            ws.cell(row=row, column=1, value=cat.category)
            ws.cell(row=row, column=2, value=float(cat.total)).number_format = '$#,##0.00'
            ws.cell(row=row, column=3, value=cat.count)
            ws.cell(row=row, column=4, value=cat.percentage / 100).number_format = '0.0%'
            row += 1

        # Create pie chart for categories
        if report.summary.by_category:
            chart = PieChart()
            chart.title = "Spending by Category"
            chart.width = 15
            chart.height = 10

            # Data for chart (using the category data we just wrote)
            data_start = row - len(report.summary.by_category[:10])
            data = Reference(ws, min_col=2, min_row=data_start,
                           max_row=row - 1)
            cats = Reference(ws, min_col=1, min_row=data_start,
                           max_row=row - 1)
            chart.add_data(data)
            chart.set_categories(cats)

            ws.add_chart(chart, "F6")

        # Top Vendors section
        row += 2
        ws.cell(row=row, column=1, value="TOP VENDORS").font = Font(bold=True, color=BRAND_GREEN, size=14)
        row += 2

        headers = ["Vendor", "Amount", "Count", "Avg Transaction", "Recurring"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, 5)

        row += 1
        for vendor in report.summary.by_vendor[:10]:
            ws.cell(row=row, column=1, value=vendor.vendor)
            ws.cell(row=row, column=2, value=float(vendor.total)).number_format = '$#,##0.00'
            ws.cell(row=row, column=3, value=vendor.count)
            ws.cell(row=row, column=4, value=float(vendor.average)).number_format = '$#,##0.00'
            ws.cell(row=row, column=5, value="Yes" if vendor.is_recurring else "No")
            row += 1

        self._auto_column_width(ws)

    def _create_transactions_sheet(self, wb: 'Workbook', report: 'Report'):
        """Create detailed transactions sheet."""
        ws = wb.create_sheet("Transactions")

        # Headers
        headers = [
            "Index", "Date", "Description", "Amount", "Category",
            "Business Type", "Review Status", "Has Receipt", "Notes", "Receipt URL"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, 1, len(headers))

        # Data rows
        for row_num, txn in enumerate(report.transactions, 2):
            ws.cell(row=row_num, column=1, value=txn.index)
            ws.cell(row=row_num, column=2, value=txn.date).number_format = 'MM/DD/YYYY' if txn.date else ''
            ws.cell(row=row_num, column=3, value=txn.description)
            ws.cell(row=row_num, column=4, value=float(txn.amount)).number_format = '$#,##0.00'
            ws.cell(row=row_num, column=5, value=txn.effective_category)
            ws.cell(row=row_num, column=6, value=txn.business_type)
            ws.cell(row=row_num, column=7, value=txn.review_status or "")
            ws.cell(row=row_num, column=8, value="Yes" if txn.has_receipt else "No")
            ws.cell(row=row_num, column=9, value=txn.notes or txn.ai_note or "")
            ws.cell(row=row_num, column=10, value=txn.effective_receipt_url)

            # Conditional formatting for receipts
            receipt_cell = ws.cell(row=row_num, column=8)
            if txn.has_receipt:
                receipt_cell.fill = PatternFill(start_color=SUCCESS_COLOR, end_color=SUCCESS_COLOR, fill_type='solid')
                receipt_cell.font = Font(color="000000")
            else:
                receipt_cell.fill = PatternFill(start_color=ERROR_COLOR, end_color=ERROR_COLOR, fill_type='solid')
                receipt_cell.font = Font(color="FFFFFF")

        # Freeze header row
        ws.freeze_panes = 'A2'

        self._auto_column_width(ws)

    def _create_category_sheet(self, wb: 'Workbook', report: 'Report'):
        """Create category breakdown sheet with charts."""
        ws = wb.create_sheet("By Category")

        # Title
        ws['A1'] = "Category Breakdown"
        ws['A1'].font = Font(bold=True, color=BRAND_GREEN, size=16)

        # Headers
        headers = ["Category", "Total Amount", "Transaction Count", "% of Total", "Avg Transaction"]
        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        # Data
        row += 1
        for cat in report.summary.by_category:
            ws.cell(row=row, column=1, value=cat.category)
            ws.cell(row=row, column=2, value=float(cat.total)).number_format = '$#,##0.00'
            ws.cell(row=row, column=3, value=cat.count)
            ws.cell(row=row, column=4, value=cat.percentage / 100).number_format = '0.0%'
            avg = float(cat.total) / cat.count if cat.count else 0
            ws.cell(row=row, column=5, value=avg).number_format = '$#,##0.00'
            row += 1

        # Add bar chart
        if report.summary.by_category:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Spending by Category"
            chart.y_axis.title = "Amount ($)"
            chart.x_axis.title = "Category"
            chart.width = 20
            chart.height = 12

            data = Reference(ws, min_col=2, min_row=3,
                           max_row=3 + len(report.summary.by_category), max_col=2)
            cats = Reference(ws, min_col=1, min_row=4,
                           max_row=3 + len(report.summary.by_category))

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4

            ws.add_chart(chart, "G3")

        self._auto_column_width(ws)

    def _create_vendor_sheet(self, wb: 'Workbook', report: 'Report'):
        """Create vendor analysis sheet."""
        ws = wb.create_sheet("By Vendor")

        # Title
        ws['A1'] = "Vendor Analysis"
        ws['A1'].font = Font(bold=True, color=BRAND_GREEN, size=16)

        # Headers
        headers = [
            "Vendor", "Total Amount", "Count", "Avg Transaction",
            "First Transaction", "Last Transaction", "Recurring", "Categories"
        ]
        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        # Data
        row += 1
        for vendor in report.summary.by_vendor:
            ws.cell(row=row, column=1, value=vendor.vendor)
            ws.cell(row=row, column=2, value=float(vendor.total)).number_format = '$#,##0.00'
            ws.cell(row=row, column=3, value=vendor.count)
            ws.cell(row=row, column=4, value=float(vendor.average)).number_format = '$#,##0.00'
            ws.cell(row=row, column=5, value=vendor.first_transaction).number_format = 'MM/DD/YYYY'
            ws.cell(row=row, column=6, value=vendor.last_transaction).number_format = 'MM/DD/YYYY'
            recurring_cell = ws.cell(row=row, column=7, value="Yes" if vendor.is_recurring else "No")
            if vendor.is_recurring:
                recurring_cell.font = Font(color=SUCCESS_COLOR)
            ws.cell(row=row, column=8, value=", ".join(vendor.categories[:3]))
            row += 1

        # Freeze header
        ws.freeze_panes = 'A4'

        self._auto_column_width(ws)

    def _create_trends_sheet(self, wb: 'Workbook', report: 'Report'):
        """Create monthly trends sheet with line chart."""
        ws = wb.create_sheet("Monthly Trends")

        # Title
        ws['A1'] = "Monthly Spending Trends"
        ws['A1'].font = Font(bold=True, color=BRAND_GREEN, size=16)

        if not report.summary.monthly_trends:
            ws['A3'] = "No trend data available"
            return

        # Headers
        headers = ["Month", "Total Amount", "Transaction Count", "Avg Transaction"]
        row = 3
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        # Data
        row += 1
        for trend in report.summary.monthly_trends:
            month_name = datetime(trend.year, trend.month, 1).strftime('%B %Y')
            ws.cell(row=row, column=1, value=month_name)
            ws.cell(row=row, column=2, value=float(trend.total)).number_format = '$#,##0.00'
            ws.cell(row=row, column=3, value=trend.count)
            avg = float(trend.total) / trend.count if trend.count else 0
            ws.cell(row=row, column=4, value=avg).number_format = '$#,##0.00'
            row += 1

        # Add line chart
        chart = LineChart()
        chart.title = "Monthly Spending"
        chart.y_axis.title = "Amount ($)"
        chart.x_axis.title = "Month"
        chart.width = 18
        chart.height = 10

        data = Reference(ws, min_col=2, min_row=3,
                        max_row=3 + len(report.summary.monthly_trends))
        cats = Reference(ws, min_col=1, min_row=4,
                        max_row=3 + len(report.summary.monthly_trends))

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "F3")

        self._auto_column_width(ws)

    def export_multi_business_report(self, reports: Dict[str, 'Report']) -> bytes:
        """
        Export multiple business type reports to single workbook.

        Args:
            reports: Dict mapping business type to Report

        Returns:
            Excel file as bytes
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl required for Excel export")

        wb = Workbook()
        self.styles_created = False
        self._create_styles(wb)

        # Create combined summary sheet
        ws = wb.active
        ws.title = "Overview"

        ws['A1'] = "Business Expense Overview"
        ws['A1'].font = Font(bold=True, color=BRAND_GREEN, size=18)

        # Overview table
        row = 3
        headers = ["Business Type", "Total Amount", "Transactions", "Match Rate", "Receipt Rate"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._apply_header_style(ws, row, 1, len(headers))

        row += 1
        grand_total = Decimal('0')
        grand_count = 0

        for bt, report in reports.items():
            if report.summary.total_transactions > 0:
                ws.cell(row=row, column=1, value=bt)
                ws.cell(row=row, column=2, value=float(report.summary.total_amount)).number_format = '$#,##0.00'
                ws.cell(row=row, column=3, value=report.summary.total_transactions)
                ws.cell(row=row, column=4, value=report.summary.match_rate / 100).number_format = '0.0%'
                ws.cell(row=row, column=5, value=report.summary.receipt_rate / 100).number_format = '0.0%'
                row += 1
                grand_total += report.summary.total_amount
                grand_count += report.summary.total_transactions

        # Grand total row
        row += 1
        ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(bold=True, color=BRAND_GREEN)
        ws.cell(row=row, column=2, value=float(grand_total)).number_format = '$#,##0.00'
        ws.cell(row=row, column=2).font = Font(bold=True, color=BRAND_GREEN)
        ws.cell(row=row, column=3, value=grand_count).font = Font(bold=True)

        self._auto_column_width(ws)

        # Create sheet for each business type
        for bt, report in reports.items():
            if report.summary.total_transactions > 0:
                # Create transactions sheet
                sheet_name = bt[:20].replace('/', '-')  # Excel sheet name limit
                ws = wb.create_sheet(sheet_name)

                # Add basic info
                ws['A1'] = f"{bt} Expenses"
                ws['A1'].font = Font(bold=True, color=BRAND_GREEN, size=16)
                ws['A2'] = f"Total: ${float(report.summary.total_amount):,.2f} ({report.summary.total_transactions} transactions)"

                # Transaction headers
                headers = ["Date", "Description", "Amount", "Category", "Status", "Receipt"]
                row = 4
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header)
                self._apply_header_style(ws, row, 1, len(headers))

                # Transaction data
                row += 1
                for txn in report.transactions:
                    ws.cell(row=row, column=1, value=txn.date).number_format = 'MM/DD/YYYY' if txn.date else ''
                    ws.cell(row=row, column=2, value=txn.description[:50])
                    ws.cell(row=row, column=3, value=float(txn.amount)).number_format = '$#,##0.00'
                    ws.cell(row=row, column=4, value=txn.effective_category)
                    ws.cell(row=row, column=5, value=txn.review_status or "")
                    ws.cell(row=row, column=6, value="Yes" if txn.has_receipt else "No")
                    row += 1

                ws.freeze_panes = 'A5'
                self._auto_column_width(ws)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()


# =============================================================================
# SINGLETON ACCESS
# =============================================================================

_exporter_instance = None

def get_excel_exporter() -> ExcelExporter:
    """Get or create Excel exporter instance."""
    global _exporter_instance
    if _exporter_instance is None:
        _exporter_instance = ExcelExporter()
    return _exporter_instance
